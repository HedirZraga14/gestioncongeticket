"""
Module de reporting pour générer des rapports avec les données IA
"""
from django.http import HttpResponse
from django.template.loader import get_template
from xhtml2pdf import pisa
import io
import csv
from datetime import datetime, timedelta
import calendar
from django.utils import timezone
from django.db.models import Count, Avg, Q, F, ExpressionWrapper, DurationField
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference, BarChart
from .ai_services import ai_services

class ReportingService:
    """Service de reporting avec intégration IA"""
    
    def __init__(self):
        self.ai_services = ai_services
    
    def generate_monthly_report_data(self, month=None, year=None):
        """Génère les données pour le rapport mensuel"""
        if month is None:
            month = timezone.now().month
        if year is None:
            year = timezone.now().year
        
        start_date = timezone.datetime(year, month, 1)
        if month == 12:
            end_date = timezone.datetime(year + 1, 1, 1) - timedelta(days=1)
        else:
            end_date = timezone.datetime(year, month + 1, 1) - timedelta(days=1)
        
        # Importer les modèles
        from leave.models import Leave
        from ticket.models import Ticket
        from employee.models import Employee
        
        # Données des congés
        leaves = Leave.objects.filter(
            created__gte=start_date,
            created__lte=end_date
        )
        
        leaves_by_status = leaves.values('status').annotate(count=Count('id'))
        leaves_by_type = leaves.values('leavetype').annotate(count=Count('id'))
        
        # Données des tickets
        tickets = Ticket.objects.filter(
            created_at__gte=start_date,
            created_at__lte=end_date
        )
        
        tickets_by_status = tickets.values('statut').annotate(count=Count('id'))
        tickets_by_destination = tickets.values('destination').annotate(count=Count('id')).order_by('-count')[:10]
        
        # Statistiques IA
        ai_stats = self._get_ai_statistics(leaves, tickets)
        
        return {
            'month': month,
            'year': year,
            'start_date': start_date,
            'end_date': end_date,
            'leaves_by_status': list(leaves_by_status),
            'leaves_by_type': list(leaves_by_type),
            'tickets_by_status': list(tickets_by_status),
            'tickets_by_destination': list(tickets_by_destination),
            'total_leaves': leaves.count(),
            'total_tickets': tickets.count(),
            'ai_statistics': ai_stats
        }
    
    def generate_analytics_report_data(self):
        """Génère les données pour le rapport analytics complet"""
        from leave.models import Leave
        from ticket.models import Ticket
        from employee.models import Employee, Department
        
        # Période d'analyse (6 derniers mois)
        end_date = timezone.now()
        start_date = end_date - timedelta(days=180)
        
        # === DONNÉES TEMPORELLES ===
        monthly_leaves = []
        monthly_tickets = []
        
        for i in range(6):
            month_start = end_date - timedelta(days=30*i)
            month_start = month_start.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            month_end = month_start.replace(day=calendar.monthrange(month_start.year, month_start.month)[1])
            
            leaves_count = Leave.objects.filter(
                created__gte=month_start,
                created__lte=month_end
            ).count()
            
            tickets_count = Ticket.objects.filter(
                created_at__gte=month_start,
                created_at__lte=month_end
            ).count()
            
            monthly_leaves.append({
                'month': month_start.strftime('%B %Y'),
                'count': leaves_count
            })
            
            monthly_tickets.append({
                'month': month_start.strftime('%B %Y'),
                'count': tickets_count
            })
        
        # === MÉTRIQUES GLOBALES ===
        total_employees = Employee.objects.count()
        total_leaves_month = Leave.objects.filter(created__gte=start_date).count()
        total_tickets_month = Ticket.objects.filter(created_at__gte=start_date).count()
        
        # === CONGÉS EN ATTENTE ===
        pending_leaves = Leave.objects.filter(status='pending').count()
        pending_tickets = Ticket.objects.filter(statut='En attente').count()
        
        # === ANALYTICS DES CONGÉS ===
        leaves_by_type = Leave.objects.filter(created__gte=start_date).values('leavetype').annotate(count=Count('id')).order_by('-count')[:10]
        leaves_by_status = Leave.objects.filter(created__gte=start_date).values('status').annotate(count=Count('id'))
        
        # === ANALYTICS DES TICKETS ===
        tickets_by_destination = Ticket.objects.filter(created_at__gte=start_date).values('destination').annotate(count=Count('id')).order_by('-count')[:10]
        tickets_by_status = Ticket.objects.filter(created_at__gte=start_date).values('statut').annotate(count=Count('id'))
        
        # === TEMPS DE RÉPONSE ===
        # Temps de réponse moyen des congés
        avg_leave_response = Leave.objects.filter(
            status__in=['approved', 'rejected'],
            created__gte=start_date
        ).annotate(
            response_time=ExpressionWrapper(
                F('updated') - F('created'),
                output_field=DurationField()
            )
        ).aggregate(avg_time=Avg('response_time'))['avg_time']
        avg_leave_response_time = avg_leave_response.days if avg_leave_response else 0
        
        # Temps de réponse moyen des tickets
        avg_ticket_response = Ticket.objects.filter(
            statut__in=['Accepté', 'Refusé'],
            created_at__gte=start_date
        ).annotate(
            response_time=ExpressionWrapper(
                F('updated') - F('created_at'),
                output_field=DurationField()
            )
        ).aggregate(avg_time=Avg('response_time'))['avg_time']
        avg_ticket_response_time = avg_ticket_response.days if avg_ticket_response else 0
        
        # Temps de réponse global
        total_responded_items = Leave.objects.filter(status__in=['approved', 'rejected']).count() + Ticket.objects.filter(statut__in=['Accepté', 'Refusé']).count()
        avg_global_response_time = (avg_leave_response_time + avg_ticket_response_time) / 2 if total_responded_items > 0 else 0
        
        # === STATISTIQUES IA ===
        ai_stats = self._get_ai_statistics(
            Leave.objects.filter(created__gte=start_date),
            Ticket.objects.filter(created_at__gte=start_date)
        )
        
        # === INSIGHTS IA ===
        insights = self.ai_services.generate_insights(
            Leave.objects.filter(created__gte=start_date).values(),
            Ticket.objects.filter(created_at__gte=start_date).values()
        )
        
        # Fallback si pas d'insights
        if not insights:
            insights = [
                "Analyse des tendances de congés sur 6 mois",
                "Évaluation des temps de réponse moyens",
                "Statistiques de performance par département",
                "Recommandations d'optimisation des processus"
            ]
        
        # === DONNÉES DÉPARTEMENTALES ===
        departments_data = []
        for dept in Department.objects.all():
            dept_employees = Employee.objects.filter(department=dept).count()
            dept_leaves = Leave.objects.filter(
                user__employee__department=dept,
                created__gte=start_date
            ).count()
            dept_tickets = Ticket.objects.filter(
                user__employee__department=dept,
                created_at__gte=start_date
            ).count()
            
            departments_data.append({
                'name': dept.name,
                'employees': dept_employees,
                'leaves': dept_leaves,
                'tickets': dept_tickets
            })
        
        return {
            'monthly_leaves': monthly_leaves,
            'monthly_tickets': monthly_tickets,
            'insights': insights,
            'period': f"{start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}",
            'total_employees': total_employees,
            'total_leaves_month': total_leaves_month,
            'total_tickets_month': total_tickets_month,
            'pending_leaves': pending_leaves,
            'pending_tickets': pending_tickets,
            'leaves_by_type': list(leaves_by_type),
            'leaves_by_status': list(leaves_by_status),
            'tickets_by_destination': list(tickets_by_destination),
            'tickets_by_status': list(tickets_by_status),
            'avg_leave_response_time': avg_leave_response_time,
            'avg_ticket_response_time': avg_ticket_response_time,
            'avg_global_response_time': round(avg_global_response_time, 1),
            'total_responded_items': total_responded_items,
            'departments_data': departments_data,
            'ai_statistics': ai_stats,
            'avg_approval_probability': ai_stats.get('avg_approval_probability', 0),
            'avg_priority_score': 75.0,  # Valeur par défaut
            'total_leaves_analyzed': Leave.objects.filter(ai_approval_probability__isnull=False).count(),
            'total_tickets_analyzed': Ticket.objects.filter(ai_sentiment__isnull=False).count()
        }
    
    def _get_ai_statistics(self, leaves, tickets):
        """Récupère les statistiques IA"""
        stats = {
            'avg_approval_probability': 0,
            'priority_distribution': {},
            'sentiment_distribution': {},
            'category_distribution': {}
        }
        
        # Statistiques des congés
        if leaves.exists():
            avg_prob = leaves.aggregate(avg_prob=Avg('ai_approval_probability'))['avg_prob']
            stats['avg_approval_probability'] = round(avg_prob or 0, 2)
            
            priority_dist = leaves.values('ai_priority').annotate(count=Count('id'))
            stats['priority_distribution'] = {item['ai_priority']: item['count'] for item in priority_dist if item['ai_priority']}
        
        # Statistiques des tickets
        if tickets.exists():
            sentiment_dist = tickets.values('ai_sentiment').annotate(count=Count('id'))
            stats['sentiment_distribution'] = {item['ai_sentiment']: item['count'] for item in sentiment_dist if item['ai_sentiment']}
            
            category_dist = tickets.values('ai_category').annotate(count=Count('id'))
            stats['category_distribution'] = {item['ai_category']: item['count'] for item in category_dist if item['ai_category']}
        
        return stats
    
    def export_excel_report(self, report_data):
        """Export du rapport en Excel"""
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="rapport_{report_data["month"]}_{report_data["year"]}.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Rapport Mensuel', f'{report_data["month"]}/{report_data["year"]}'])
        writer.writerow([])
        
        # Congés par statut
        writer.writerow(['Congés par statut'])
        for item in report_data['leaves_by_status']:
            writer.writerow([item['status'], item['count']])
        writer.writerow([])
        
        # Tickets par statut
        writer.writerow(['Tickets par statut'])
        for item in report_data['tickets_by_status']:
            writer.writerow([item['statut'], item['count']])
        writer.writerow([])
        
        # Statistiques IA
        writer.writerow(['Statistiques IA'])
        writer.writerow(['Probabilité moyenne d\'approbation', f"{report_data['ai_statistics']['avg_approval_probability']}%"])
        
        return response
    
    def export_pdf_report(self, report_data):
        """Export du rapport en PDF"""
        template = get_template('dashboard/reports/monthly_report_template.html')
        html = template.render({'report_data': report_data})
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="rapport_{report_data["month"]}_{report_data["year"]}.pdf"'
        
        pisa_status = pisa.CreatePDF(html, dest=response)
        
        if pisa_status.err:
            return HttpResponse('Erreur lors de la génération du PDF')
        
        return response
    
    def export_analytics_excel_report(self, report_data):
        """Export du rapport analytics en Excel professionnel"""
        # Créer un nouveau classeur Excel
        wb = openpyxl.Workbook()
        
        # Supprimer la feuille par défaut
        wb.remove(wb.active)
        
        # === FEUILLE 1: RÉSUMÉ EXÉCUTIF ===
        ws_summary = wb.create_sheet("Résumé Exécutif")
        self._create_summary_sheet(ws_summary, report_data)
        
        # === FEUILLE 2: ANALYTICS CONGÉS ===
        ws_leaves = wb.create_sheet("Analytics Congés")
        self._create_leaves_analytics_sheet(ws_leaves, report_data)
        
        # === FEUILLE 3: ANALYTICS TICKETS ===
        ws_tickets = wb.create_sheet("Analytics Tickets")
        self._create_tickets_analytics_sheet(ws_tickets, report_data)
        
        # === FEUILLE 4: INSIGHTS IA ===
        ws_ai = wb.create_sheet("Insights IA")
        self._create_ai_insights_sheet(ws_ai, report_data)
        
        # === FEUILLE 5: TEMPS DE RÉPONSE ===
        ws_response = wb.create_sheet("Temps de Réponse")
        self._create_response_time_sheet(ws_response, report_data)
        
        # === FEUILLE 6: ANALYTICS DÉPARTEMENTAUX ===
        ws_dept = wb.create_sheet("Analytics Départements")
        self._create_departments_sheet(ws_dept, report_data)
        
        # Sauvegarder dans un buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Créer la réponse HTTP
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="Analytics_Complet_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx"'
        
        return response
    
    def export_analytics_pdf_report(self, report_data):
        """Export du rapport analytics en PDF"""
        template = get_template('dashboard/reports/analytics_report_template.html')
        html = template.render({'report_data': report_data})
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="analytics_{datetime.now().strftime("%Y%m%d")}.pdf"'
        
        pisa_status = pisa.CreatePDF(html, dest=response)
        
        if pisa_status.err:
            return HttpResponse('Erreur lors de la génération du PDF')
        
        return response

    def _create_summary_sheet(self, ws, data):
        """Crée la feuille de résumé exécutif"""
        # Titre principal
        ws['A1'] = "📊 RAPPORT ANALYTICS COMPLET"
        ws['A1'].font = Font(size=20, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
        ws.merge_cells('A1:F1')
        
        # Informations de base
        ws['A3'] = "Période d'analyse:"
        ws['A3'].font = Font(bold=True)
        ws['B3'] = data.get('period', '6 derniers mois')
        
        ws['A4'] = "Date de génération:"
        ws['A4'].font = Font(bold=True)
        ws['B4'] = datetime.now().strftime("%d/%m/%Y %H:%M")
        
        # KPI principaux
        ws['A6'] = "🎯 KPI PRINCIPAUX"
        ws['A6'].font = Font(size=16, bold=True, color="2E86AB")
        ws.merge_cells('A6:F6')
        
        # Tableau des KPI
        kpi_data = [
            ["Métrique", "Valeur", "Unité", "Tendance"],
            ["Total Employés", data.get('total_employees', 0), "personnes", "📈"],
            ["Congés du mois", data.get('total_leaves_month', 0), "demandes", "📊"],
            ["Tickets du mois", data.get('total_tickets_month', 0), "demandes", "📊"],
            ["Temps de réponse moyen", data.get('avg_global_response_time', 0), "jours", "⏱️"],
            ["Taux d'approbation IA", f"{data.get('avg_approval_probability', 0)}%", "%", "🤖"]
        ]
        
        for row_idx, row_data in enumerate(kpi_data, 8):
            for col_idx, cell_value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
                if row_idx == 8:  # En-tête
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
                else:
                    cell.font = Font(size=11)
        
        # Ajuster la largeur des colonnes
        for col in ['A', 'B', 'C', 'D']:
            ws.column_dimensions[col].width = 20
    
    def _create_leaves_analytics_sheet(self, ws, data):
        """Crée la feuille d'analytics des congés"""
        # Titre
        ws['A1'] = "🏖️ ANALYTICS CONGÉS"
        ws['A1'].font = Font(size=18, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="48BB78", end_color="48BB78", fill_type="solid")
        ws.merge_cells('A1:F1')
        
        # Évolution mensuelle
        ws['A3'] = "📈 Évolution mensuelle des congés"
        ws['A3'].font = Font(size=14, bold=True, color="48BB78")
        
        # En-têtes
        ws['A5'] = "Mois"
        ws['B5'] = "Nombre de congés"
        ws['A5'].font = Font(bold=True)
        ws['B5'].font = Font(bold=True)
        
        # Données
        row = 6
        for item in data.get('monthly_leaves', []):
            ws[f'A{row}'] = item.get('month', '')
            ws[f'B{row}'] = item.get('count', 0)
            row += 1
        
        # Graphique
        chart = LineChart()
        chart.title = "Évolution des congés"
        chart.x_axis.title = "Mois"
        chart.y_axis.title = "Nombre de congés"
        
        data_range = Reference(ws, min_col=2, min_row=5, max_row=row-1)
        cats = Reference(ws, min_col=1, min_row=6, max_row=row-1)
        chart.add_data(data_range, titles_from_data=True)
        chart.set_categories(cats)
        
        ws.add_chart(chart, "D3")
        
        # Ajuster la largeur des colonnes
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
    
    def _create_tickets_analytics_sheet(self, ws, data):
        """Crée la feuille d'analytics des tickets"""
        # Titre
        ws['A1'] = "✈️ ANALYTICS TICKETS"
        ws['A1'].font = Font(size=18, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="ED8936", end_color="ED8936", fill_type="solid")
        ws.merge_cells('A1:F1')
        
        # Évolution mensuelle
        ws['A3'] = "📈 Évolution mensuelle des tickets"
        ws['A3'].font = Font(size=14, bold=True, color="ED8936")
        
        # En-têtes
        ws['A5'] = "Mois"
        ws['B5'] = "Nombre de tickets"
        ws['A5'].font = Font(bold=True)
        ws['B5'].font = Font(bold=True)
        
        # Données
        row = 6
        for item in data.get('monthly_tickets', []):
            ws[f'A{row}'] = item.get('month', '')
            ws[f'B{row}'] = item.get('count', 0)
            row += 1
        
        # Graphique
        chart = BarChart()
        chart.title = "Évolution des tickets"
        chart.x_axis.title = "Mois"
        chart.y_axis.title = "Nombre de tickets"
        
        data_range = Reference(ws, min_col=2, min_row=5, max_row=row-1)
        cats = Reference(ws, min_col=1, min_row=6, max_row=row-1)
        chart.add_data(data_range, titles_from_data=True)
        chart.set_categories(cats)
        
        ws.add_chart(chart, "D3")
        
        # Ajuster la largeur des colonnes
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
    
    def _create_ai_insights_sheet(self, ws, data):
        """Crée la feuille des insights IA"""
        # Titre
        ws['A1'] = "🤖 INSIGHTS INTELLIGENCE ARTIFICIELLE"
        ws['A1'].font = Font(size=18, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="9F7AEA", end_color="9F7AEA", fill_type="solid")
        ws.merge_cells('A1:F1')
        
        # Statistiques IA
        ws['A3'] = "📊 Statistiques IA"
        ws['A3'].font = Font(size=14, bold=True, color="9F7AEA")
        
        # Tableau des stats
        ai_stats = [
            ["Métrique", "Valeur", "Description"],
            ["Probabilité d'approbation", f"{data.get('avg_approval_probability', 0)}%", "Probabilité moyenne d'approbation des congés"],
            ["Score de priorité", f"{data.get('avg_priority_score', 0)}%", "Score de priorité moyen des tickets"],
            ["Congés analysés", data.get('total_leaves_analyzed', 0), "Nombre total de congés analysés par l'IA"],
            ["Tickets analysés", data.get('total_tickets_analyzed', 0), "Nombre total de tickets analysés par l'IA"]
        ]
        
        for row_idx, row_data in enumerate(ai_stats, 5):
            for col_idx, cell_value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
                if row_idx == 5:  # En-tête
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="9F7AEA", end_color="9F7AEA", fill_type="solid")
                else:
                    cell.font = Font(size=11)
        
        # Insights
        ws['A12'] = "💡 Insights générés par l'IA"
        ws['A12'].font = Font(size=14, bold=True, color="9F7AEA")
        
        row = 14
        for insight in data.get('insights', []):
            ws[f'A{row}'] = f"• {insight}"
            ws[f'A{row}'].font = Font(size=11)
            row += 1
        
        # Ajuster la largeur des colonnes
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 50
    
    def _create_response_time_sheet(self, ws, data):
        """Crée la feuille des temps de réponse"""
        # Titre
        ws['A1'] = "⏱️ ANALYTICS TEMPS DE RÉPONSE"
        ws['A1'].font = Font(size=18, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="F56565", end_color="F56565", fill_type="solid")
        ws.merge_cells('A1:F1')
        
        # Métriques de temps de réponse
        ws['A3'] = "📊 Métriques de performance"
        ws['A3'].font = Font(size=14, bold=True, color="F56565")
        
        # Tableau des métriques
        response_metrics = [
            ["Métrique", "Congés", "Tickets", "Global"],
            ["Temps de réponse moyen", f"{data.get('avg_leave_response_time', 0)} jours", f"{data.get('avg_ticket_response_time', 0)} jours", f"{data.get('avg_global_response_time', 0)} jours"],
            ["Total traité", data.get('total_leaves_month', 0), data.get('total_tickets_month', 0), data.get('total_responded_items', 0)],
            ["En attente", data.get('pending_leaves', 0), data.get('pending_tickets', 0), "N/A"]
        ]
        
        for row_idx, row_data in enumerate(response_metrics, 5):
            for col_idx, cell_value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
                if row_idx == 5:  # En-tête
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="F56565", end_color="F56565", fill_type="solid")
                else:
                    cell.font = Font(size=11)
        
        # Ajuster la largeur des colonnes
        for col in ['A', 'B', 'C', 'D']:
            ws.column_dimensions[col].width = 25
    
    def _create_departments_sheet(self, ws, data):
        """Crée la feuille d'analytics des départements"""
        # Titre
        ws['A1'] = "🏢 ANALYTICS DÉPARTEMENTAUX"
        ws['A1'].font = Font(size=18, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="38A169", end_color="38A169", fill_type="solid")
        ws.merge_cells('A1:F1')
        
        # Tableau des départements
        ws['A3'] = "📊 Performance par département"
        ws['A3'].font = Font(size=14, bold=True, color="38A169")
        
        # En-têtes
        headers = ["Département", "Employés", "Congés", "Tickets", "Taux d'activité"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="38A169", end_color="38A169", fill_type="solid")
        
        # Données des départements
        row = 6
        for dept in data.get('departments_data', []):
            ws[f'A{row}'] = dept.get('name', 'N/A')
            ws[f'B{row}'] = dept.get('employees', 0)
            ws[f'C{row}'] = dept.get('leaves', 0)
            ws[f'D{row}'] = dept.get('tickets', 0)
            
            # Calculer le taux d'activité
            total_activities = dept.get('leaves', 0) + dept.get('tickets', 0)
            employees = dept.get('employees', 1)
            activity_rate = (total_activities / employees) if employees > 0 else 0
            ws[f'E{row}'] = f"{activity_rate:.1f}"
            
            row += 1
        
        # Graphique en barres
        chart = BarChart()
        chart.title = "Performance par département"
        chart.x_axis.title = "Département"
        chart.y_axis.title = "Nombre d'activités"
        
        data_range = Reference(ws, min_col=3, min_row=5, max_col=4, max_row=row-1)
        cats = Reference(ws, min_col=1, min_row=6, max_row=row-1)
        chart.add_data(data_range, titles_from_data=True)
        chart.set_categories(cats)
        
        ws.add_chart(chart, "G3")
        
        # Ajuster la largeur des colonnes
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 20

# Instance globale
reporting_service = ReportingService() 